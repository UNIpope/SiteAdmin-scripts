import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.styles import PatternFill
import sys, re

# read in rotations.csv + cleanup
df = pd.read_csv(r"siteadmin\timesheet\8august\aug8rotation.csv")
df = df.fillna("I")
df = df.apply(lambda x: x.astype(str).str.lower().str.strip())

# read in excel wkbk 
mainwb = r"siteadmin\timesheet\8august\CMI DE1 Timesheet - WW 35 27-08-20.xlsx"
workbook = xl.load_workbook(filename=mainwb)
sheet = workbook.active

#date handler
date = sheet["O8"].value
date = date.split()[2]
weekend = date.split("/")[0]
month = date.split("/")[1]
print("week, month:", weekend, ", ", month)

# get start of week and generate week dates
weekst = str(int(weekend) - 6)
weekhead =  ["Name"] + list( map( str, range( int(weekst), int(weekend) + 1 )))

# filter rotationcsv to desired dates
filtereddf = df[weekhead]
workingwk = ["Name","Friday","Saturday","Sunday","Monday","Tuesday","Wednesday", "Thursday"] # col name label
filtereddf.columns = workingwk

print(filtereddf)

# generate lookup table 
name_cell = {}
names  = sheet["D"]
for cell in names[14:]:
    if cell.value:
        name = str.lower(cell.value).strip().split(",")
        name = " ".join(name[::-1]).strip()
        name_cell[name] =  cell.row

print(name_cell)

#fill logic
dayc = {"Friday":("E","F"), "Saturday":("G","H"), "Sunday":("I","J"), 
        "Monday":("K","L"), "Tuesday":("M","N"), "Wednesday":("O","P"), 
        "Thursday":("Q","R")}

err = [] 
for _, row in filtereddf.iterrows():
    ckey = row["Name"].lower()
    
    try:
        rw = name_cell[ckey]
        print()
        print(ckey, rw)

        for day in dayc.keys():
            cords = ["",""]
            
            cords[0] = dayc[day][0] + str(rw)
            cords[1] = dayc[day][1] + str(rw)

            print(day, cords)

            if day == "Saturday" or day == "Sunday":
                pass
            else:
                if row[day] == "h":
                    intime = "07:30:00"
                    outime = "12:00:00"

                    sheet[cords[0]] = intime
                    sheet[cords[1]] = outime

                    c = PatternFill(fgColor="FFFF00", fill_type = "solid")
                    sheet[cords[0]].fill = c
                    sheet[cords[1]].fill = c

                elif row[day] == "i":
                    intime = "07:30:00"
                    outime = "18:00:00"

                    sheet[cords[0]] = intime
                    sheet[cords[1]] = outime
                
                elif row[day] == "v":
                    c = PatternFill(fgColor="FF0000", fill_type = "solid")
                    sheet[cords[0]].fill = c
                    sheet[cords[1]].fill = c
                
                elif row[day] == "l":
                    c = PatternFill(fgColor="FF681F", fill_type = "solid")
                    sheet[cords[0]].fill = c
                    sheet[cords[1]].fill = c
                
                elif row[day] == "IRE":
                    c = PatternFill(fgColor="90EE90", fill_type = "solid")
                    sheet[cords[0]].fill = c
                    sheet[cords[1]].fill = c

                else:
                    print(day, row[day])

    except:
        err.append(ckey)

print(err)
workbook.save(mainwb)

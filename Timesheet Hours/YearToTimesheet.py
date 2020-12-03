import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import sys, calendar
from collections import OrderedDict

# read in rotations.csv + cleanup
df = pd.read_csv(r"C:\Users\jduggan\Documents\A hole year ivana\merge.csv")
df = df.fillna("not_emp")
df = df.apply(lambda x: x.astype(str).str.lower().str.strip())

# get working week ranges
weeks = []
for col_name in df.columns:
    if "Thu" in col_name:
        week_end = df.columns.get_loc(col_name)
        # handel starting of year
        if week_end - 6 > 0:
            # get day past the week other sides because of flights
            weeks.append(range(week_end - 7 , week_end+2))
        else:
            weeks.append(range(1, week_end+2))

#each week
for week in weeks:
    weeknum = str(weeks.index(week) + 1)

    # format week and add name to begining
    week = list(week)
    week.insert(0,0)

    # apply week to df
    try:
        wkweek = df.iloc[range(0,df.shape[0]), week]
    except IndexError:
        break
    
    # filter out the unemployed, finished and starting people
    for index, row in wkweek.iterrows():
        r = list(row)[2:-1]

        cnotemp = r.count("not_emp")
        if len(r) == cnotemp:
            wkweek = wkweek.drop(index)

        cnotemp = r.count("fn")
        if len(r) == cnotemp:
            wkweek = wkweek.drop(index)

        cnotemp = r.count("st")
        if len(r) == cnotemp:
            wkweek = wkweek.drop(index)


    path = r"C:\Users\jduggan\Documents\A hole year ivana\\"
    workbook = xl.load_workbook(filename=path+"FRA15 Timesheet - WW demo.xlsx")
    sheet = workbook.active

    # month and date 
    enddates = wkweek.columns[-2]
    enddates = enddates.split(":")

    monthlookup = {month.lower(): index for index, month in enumerate(calendar.month_abbr) if month}# create a month to num lookup dict
    monthnum = str(monthlookup[enddates[2][:3].lower()])# lookup month 
    datenum = enddates[1]

    #Fill date and week num
    sheet["O8"] = "Week Ending: " + datenum + "/" + monthnum + "/20"
    sheet["O10"] = "Week Number: " + weeknum

    # fill in person row-->excel
    # define colour
    yellow = PatternFill(fgColor="FFF000", fill_type = "solid")
    redaccent2 = PatternFill(fgColor="FF9999", fill_type = "solid")
    grey = PatternFill(fgColor="c0c0c0", fill_type = "solid")
    red = PatternFill(fgColor="FF0000", fill_type = "solid")
    sgreen = PatternFill(fgColor="90EE90", fill_type = "solid")
    bgreen = PatternFill(fgColor="00FF7F", fill_type = "solid")
    salmon = PatternFill(fgColor="FA8072", fill_type = "solid")
    

    # define boarders, fonts for names / h-s / sap
    generalst = NamedStyle(name="generalst")
    generalst.border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

    generalst.font = Font(name='Arial', size=18, bold=True,
                italic=False, vertAlign=None,
                underline='none', strike=False,
                color='000000')  

    generalst.alignment = Alignment(horizontal='center',
                vertical='center',text_rotation=0,
                wrap_text=False,shrink_to_fit=False, 
                indent=0)

    workbook.add_named_style(generalst)

    # define boarders, fonts for days
    dayst = NamedStyle(name="dayst")
    dayst.border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    dayst.font = Font(name='Arial', size=18, bold=True,
                italic=False, vertAlign=None,
                underline='none', strike=False,
                color='000000')  

    dayst.alignment = Alignment(horizontal='center',
                vertical='center',text_rotation=0,
                wrap_text=False,shrink_to_fit=False, 
                indent=0)

    workbook.add_named_style(dayst)


    st = 13
    for index, row in wkweek.iterrows():
        st = st + 2
        rowdata = row.to_dict(into=OrderedDict)

        # set dimention of row
        rd = sheet.row_dimensions[st] # get dim
        rd.height = 46.5

        #insert name
        sheet["D"+str(st)] = rowdata["Name"].title()
        sheet["D"+str(st)].style = "generalst"
        sheet["D"+str(st)].fill = yellow

        # fill sap style
        sheet["C"+str(st)].style = "generalst"
        sheet["C"+str(st)].fill = yellow

        # fill H/S style
        sheet["B"+str(st)] = "S"
        sheet["B"+str(st)].style = "generalst"
        sheet["B"+str(st)].fill = redaccent2

        #fill weekends
        for i in ["G","H","I","J"]:
            sheet[i+str(st)].style = "dayst"
            sheet[i+str(st)].fill = grey

            sheet[i+str(st-1)].style = "dayst"
            sheet[i+str(st-1)].fill = grey
            sheet[i+str(st-1)].number_format = "0.00"

        if st-1 != 14:
            sheet["H"+str(st-1)] = 0.00
            sheet["J"+str(st-1)] = 0.00
        
        # fill wk week data
        daydict = {"Fri":("E","F"), "Sat":("G","H"), "Sun":("I","J"), 
                    "Mon":("K","L"), "Tue":("M","N"), "Wed":("O","P"), 
                    "Thu":("Q","R")}

        rowdatals = list(rowdata.items())[1:] # drop name

        # get most occoring special day 
        specialdayvals = [x[1] for x in rowdatals if (x[1]!="in" and x[1]!="h")]
        if specialdayvals:
            specialday = max(set(specialdayvals), key = specialdayvals.count)
        else: specialday = None

        totalhrs = 0
        for daykey, val in rowdatals:
            day = daykey.split(":")[0].strip()
            daydate = daykey.split(":")[1].strip()

            # get valid keys
            vdaykeys = [x[0] for x in rowdatals][1:-1]
            if daykey in vdaykeys:
                # generate cords to write too
                cords = ["",""]
                cords[0] = daydict[day][0] + str(st)
                cords[1] = daydict[day][1] + str(st)

                ncords = ["",""]
                ncords[0] = daydict[day][0] + str(st+1)
                ncords[1] = daydict[day][1] + str(st+1)

                if day == "Sat" or day == "Sun":
                    pass
                else:
                    if val == "h": # flights
                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = yellow
                        sheet[cords[1]].fill = yellow

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00                  


                    elif val == "v": # vacation days
                        if specialday == "v":
                            sheet["S"+str(st)].value = "Vacation"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = red

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = red
                        sheet[cords[1]].fill = red

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 
                    
                    elif val == "o": # other flights
                        if specialday == "o":
                            sheet["S"+str(st)].value = "Other flights"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = yellow

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = yellow
                        sheet[cords[1]].fill = yellow

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 

                    elif val == "in": # working days
                        totalhrs = totalhrs + 10
                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"

                        sheet[cords[0]].value = "07:30:00"
                        sheet[cords[1]].value = "18:00:00"

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 10
                    
                    elif val == "l": # lieu days
                        if specialday == "l":
                            sheet["S"+str(st)].value = "Lieu"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = salmon

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = salmon
                        sheet[cords[1]].fill = salmon

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 

                    elif val == "s": # sick days
                        if specialday == "s":
                            sheet["S"+str(st)].value = "Sick"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = sgreen

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = sgreen
                        sheet[cords[1]].fill = sgreen

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 
                    
                    elif val == "ire": # bank holidays
                        if specialday == "s":
                            sheet["S"+str(st)].value = "Ire bank"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = bgreen

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = bgreen
                        sheet[cords[1]].fill = bgreen

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 

                    elif val == "ger": # bank holidays
                        if specialday == "s":
                            sheet["S"+str(st)].value = "Ger Bank"  
                            sheet["S"+str(st)].style = "dayst"
                            sheet["S"+str(st)].fill = bgreen

                        sheet[cords[0]].style = "dayst"
                        sheet[cords[1]].style = "dayst"
                        sheet[cords[0]].fill = bgreen
                        sheet[cords[1]].fill = bgreen

                        sheet[ncords[0]].style = "dayst"
                        sheet[ncords[1]].style = "dayst"
                        sheet[ncords[1]].number_format = "0.00"
                        sheet[ncords[1]].value = 0.00 

        
        # handel flight half times fill
        changels = np.array([a[1] for a in rowdatals])
        changels = np.where(np.roll(changels,1)!=changels)[0][1:]
        
        for changeindex in changels:
            if rowdatals[changeindex-1][1] == "h" and rowdatals[changeindex][1] == "in" and int(datenum) != changeindex:
                day = rowdatals[changeindex][0].split(":")[0]
                flights = daydict.get(day, False) # lookup cords to excel
                if(day == "Sat" or day =="Sun"):break

                sheet[flights[0]+str(st)].value = "12:30:00"
                sheet[flights[1]+str(st)].value = "18:00:00"

                sheet[flights[0]+str(st)].style = "dayst"
                sheet[flights[1]+str(st)].style = "dayst"
                sheet[flights[0]+str(st)].fill = yellow
                sheet[flights[1]+str(st)].fill = yellow

                totalhrs = totalhrs + 4.5 - 10

                sheet[flights[1]+str(st+1)].style = "dayst"
                sheet[flights[1]+str(st+1)].number_format = "0.00"
                sheet[flights[1]+str(st+1)].value = 4.5

            
            if rowdatals[changeindex -1][1] == "in" and rowdatals[changeindex][1] == "h":
                day = rowdatals[changeindex-1][0].split(":")[0]
                flights = daydict.get(day, False) # lookup cords to excel
                if(day == "Sat" or day =="Sun"):break

                sheet[flights[0]+str(st)].value = "07:30:00"
                sheet[flights[1]+str(st)].value = "12:00:00"

                sheet[flights[0]+str(st)].style = "dayst"
                sheet[flights[1]+str(st)].style = "dayst"
                sheet[flights[0]+str(st)].fill = yellow
                sheet[flights[1]+str(st)].fill = yellow

                if (int(datenum) >= int(daydate)) and (int(datenum) - 6 <= int(daydate)):
                    totalhrs = totalhrs + 4.5

                    sheet[flights[1]+str(st+1)].style = "dayst"
                    sheet[flights[1]+str(st+1)].number_format = "0.00"
                    sheet[flights[1]+str(st+1)].value = 4.5
        
        # write total hrs worked to sheet
        sheet["S"+str(st+1)].value = totalhrs
        sheet["S"+str(st+1)].style = "dayst"
 
    #save excel
    enddate =  monthnum+ "-" + datenum
    workbook.save(path + "FRA15 Timesheet - WW "+ weeknum + " 20-"+ enddate +".xlsx")